"""Banker-convention cell styling helpers for the workbook writer.

Encapsulates the house rules from docs/WORKBOOK_SPEC.md so the writer stays
declarative:

* **Blue font** = a hard input (only on Assumptions + the two Cover inputs).
* **Black font** = a live formula (everywhere else).
* Number formats: currency ``#,##0;(#,##0)``, percent ``0.0%``, per-share
  ``0.00``, multiple ``0.0x``.

The two cell-writing helpers are the enforcement point: ``input_cell`` writes a
blue number, ``formula_cell`` writes a black formula string (must start with
``=``). Using them consistently is what makes the "formulas-not-values" audit
pass.
"""

from __future__ import annotations

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Colors (hex, banker convention) ---
BLUE = "0000CC"  # hard inputs
BLACK = "000000"  # formulas
WHITE = "FFFFFF"
HEADER_FILL = "1F3864"  # dark navy section headers
SUBHEADER_FILL = "D9E1F2"  # light blue band

# --- Number formats ---
FMT_CURRENCY = "#,##0;(#,##0)"
FMT_PERCENT = "0.0%"
FMT_PERSHARE = "0.00"
FMT_MULTIPLE = "0.0x"
FMT_SHARES = "#,##0"
FMT_YEAR = "0"

_THIN = Side(style="thin", color="BFBFBF")
BORDER_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOP = Side(style="thin", color="000000")
BORDER_TOP = Border(top=_TOP)


def input_cell(ws, coord: str, value, fmt: str = FMT_CURRENCY) -> Cell:
    """Write a hard INPUT: blue font, numeric value (never a formula)."""
    cell = ws[coord]
    cell.value = value
    cell.font = Font(color=BLUE, name="Calibri", size=11)
    cell.number_format = fmt
    return cell


def formula_cell(ws, coord: str, formula: str, fmt: str = FMT_CURRENCY) -> Cell:
    """Write a live FORMULA: black font. ``formula`` must start with '='."""
    if not formula.startswith("="):
        raise ValueError(f"formula_cell requires a formula starting with '=': {formula!r}")
    cell = ws[coord]
    cell.value = formula
    cell.font = Font(color=BLACK, name="Calibri", size=11)
    cell.number_format = fmt
    return cell


def value_cell(ws, coord: str, value, fmt: str = FMT_CURRENCY) -> Cell:
    """Write a historical reported VALUE (the XBRL tie-out surface).

    Black font (it is not a blue input), plain number. Used only for historical
    statement lines sourced from SEC facts, per the spec.
    """
    cell = ws[coord]
    cell.value = value
    cell.font = Font(color=BLACK, name="Calibri", size=11)
    cell.number_format = fmt
    return cell


def label_cell(ws, coord: str, text: str, *, bold: bool = False, indent: int = 0) -> Cell:
    cell = ws[coord]
    cell.value = text
    cell.font = Font(bold=bold, name="Calibri", size=11)
    cell.alignment = Alignment(indent=indent)
    return cell


def section_header(ws, coord: str, text: str) -> Cell:
    cell = ws[coord]
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, name="Calibri", size=12)
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    return cell


def subheader(ws, coord: str, text: str) -> Cell:
    cell = ws[coord]
    cell.value = text
    cell.font = Font(bold=True, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
    return cell
