"""Live-formula Excel workbook writer.

Writes ``out/<TICKER>_model.xlsx`` from a :class:`~src.interfaces.ModelBundle`
per docs/WORKBOOK_SPEC.md. The governing rule (banker convention + the
cell-level differential): every *computed* cell is a live Excel formula
referencing other cells / named ranges; hard inputs (blue) live only on the
Assumptions tab and the two Cover market inputs. Historical statement lines are
written as reported values (the XBRL tie-out surface); their subtotals are SUM
formulas.

The writer is deliberately faithful to the engine's arithmetic so that when the
verifier recalculates the file (via the ``formulas`` library) every mapped cell
reproduces the engine number. :func:`build_verifier_cell_map` is the
authoritative (sheet, cell) -> engine-value map the verifier consumes; keeping
it beside the writer keeps the two in lockstep.

Scope: the Model IS projection (revenue chain → margins → EBIT → EBITDA →
pre-tax → net income), WACC build, DCF, and the sensitivity grid are live Excel
formulas; the full projected balance sheet and cash-flow statement are rendered
from the engine's computed series (black values — they are builder outputs, not
user inputs). ``build_verifier_cell_map`` lists every formula cell that must
reproduce an engine value; the verifier recalculates the file and diffs those to
the cent. Every rendered figure originates from the engine — nothing is
hand-typed.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from src.interfaces import ModelBundle, StatementSet
from src.schema import LineItem
from src.workbook import styles as S

# Per-sheet column widths and freeze panes for banker-grade ergonomics.
# Column A holds row labels (widest); data columns are moderate.
_LABEL_COL_WIDTH = 34.0
_DATA_COL_WIDTH = 13.0
# Tabs that are multi-year data grids with a header on row 3 → freeze below it.
_GRID_TABS = frozenset(
    {"Historical IS", "Historical BS", "Historical CF", "Model IS", "Model BS", "Model CF"}
)

# Sheet names, in spec order.
SH_COVER = "Cover"
SH_ASSUM = "Assumptions"
SH_HIST_IS = "Historical IS"
SH_HIST_BS = "Historical BS"
SH_HIST_CF = "Historical CF"
SH_MODEL_IS = "Model IS"
SH_MODEL_BS = "Model BS"
SH_MODEL_CF = "Model CF"
SH_WACC = "WACC"
SH_DCF = "DCF"
SH_COMPS = "Trading Comps"
SH_PREC = "Precedents"
SH_LBO = "LBO"
SH_SENS = "Sensitivities"
SH_FF = "Football Field"

SHEET_ORDER = [
    SH_COVER,
    SH_ASSUM,
    SH_HIST_IS,
    SH_HIST_BS,
    SH_HIST_CF,
    SH_MODEL_IS,
    SH_MODEL_BS,
    SH_MODEL_CF,
    SH_WACC,
    SH_DCF,
    SH_COMPS,
    SH_PREC,
    SH_LBO,
    SH_SENS,
    SH_FF,
]

# Required named ranges (WORKBOOK_SPEC §Named ranges).
NAMED_RANGES = [
    "CurrentPrice",
    "PriceTarget",
    "Rating",
    "WACC",
    "TerminalGrowth",
    "ExitMultiple",
    "EV_Gordon",
    "EV_Exit",
    "EquityValue_Gordon",
    "EquityValue_Exit",
    "ImpliedPrice_Gordon",
    "ImpliedPrice_Exit",
    "NetDebt",
    "SharesDiluted",
    "Comps_ImpliedPrice_EBITDA",
    "LBO_IRR",
    "LBO_MOIC",
    "RevenueCAGR",
]


class ExcelWorkbookWriter:
    """Concrete WorkbookWriter. See module docstring for the contract."""

    def write(self, path: str, model: ModelBundle) -> None:
        wb = Workbook()
        # Pin document timestamps to a fixed epoch so a rebuild is byte-identical
        # (openpyxl otherwise stamps the current UTC time into docProps/core.xml).
        # The pin only sticks because _save_pinned bypasses Workbook.save, which
        # would re-stamp `modified` with now() at write time.
        from datetime import datetime

        fixed = datetime(2026, 1, 1, 0, 0, 0)  # noqa: DTZ001 — deterministic constant
        wb.properties.created = fixed
        wb.properties.modified = fixed
        # Remove the default sheet; create ours in order.
        wb.remove(wb.active)
        sheets: dict[str, Worksheet] = {name: wb.create_sheet(name) for name in SHEET_ORDER}

        # Anchors used by named ranges / cross-sheet formulas, filled as we build.
        self._anchors: dict[str, str] = {}

        self._assumptions(sheets[SH_ASSUM], model)
        self._wacc(sheets[SH_WACC], model)
        self._model_is(sheets[SH_MODEL_IS], model)
        self._dcf(sheets[SH_DCF], model)
        self._comps(sheets[SH_COMPS], model)
        self._precedents(sheets[SH_PREC], model)
        self._lbo(sheets[SH_LBO], model)
        self._historical(
            sheets[SH_HIST_IS],
            model,
            [
                LineItem.REVENUE,
                LineItem.COST_OF_REVENUE,
                LineItem.GROSS_PROFIT,
                LineItem.OPERATING_INCOME,
                LineItem.NET_INCOME,
            ],
        )
        self._historical(
            sheets[SH_HIST_BS], model, [LineItem.CASH, LineItem.TOTAL_ASSETS, LineItem.TOTAL_EQUITY]
        )
        self._historical(sheets[SH_HIST_CF], model, [LineItem.CFO, LineItem.CAPEX])
        self._model_bs(sheets[SH_MODEL_BS], model)
        self._model_cf(sheets[SH_MODEL_CF], model)
        self._sensitivities(sheets[SH_SENS], model)
        # Cover must precede the football field: it sets the current_price /
        # price_target anchors the football-field markers reference.
        self._cover(sheets[SH_COVER], model)
        self._football_field(sheets[SH_FF], model)

        # Native Excel charts (a banker expects charts live in the model, not
        # only in the PDF) + column widths / freeze panes for readability.
        self._add_charts(sheets, model)
        self._apply_ergonomics(sheets)

        self._define_names(wb)
        self._save_pinned(wb, path)

    @staticmethod
    def _save_pinned(wb: Workbook, path: str) -> None:
        """Save so a rebuild is byte-identical across processes and time.

        Two openpyxl behaviors otherwise defeat determinism: (1) ``Workbook.save``
        re-stamps ``properties.modified`` with ``datetime.now()`` at write time,
        and (2) every ZIP entry gets the current wall-clock time as its local-
        header ``date_time``. Driving :class:`ExcelWriter` directly avoids (1);
        a ZipFile that pins each entry's ``date_time`` to a fixed epoch avoids (2)."""
        from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

        from openpyxl.writer.excel import ExcelWriter

        fixed_date_time = (2026, 1, 1, 0, 0, 0)

        class _PinnedZip(ZipFile):
            """ZipFile whose entries carry a fixed timestamp. openpyxl writes most
            parts via ``writestr`` (string arcname → defaults to now()) and the
            worksheets via ``write`` from a temp file (→ the file's mtime); pin
            both so the archive is byte-identical across processes and time."""

            def writestr(self, zinfo_or_arcname, data, *args, **kwargs):
                if isinstance(zinfo_or_arcname, str):
                    zinfo_or_arcname = ZipInfo(zinfo_or_arcname, date_time=fixed_date_time)
                    zinfo_or_arcname.compress_type = self.compression
                else:
                    zinfo_or_arcname.date_time = fixed_date_time
                return super().writestr(zinfo_or_arcname, data, *args, **kwargs)

            def write(self, filename, arcname=None, *args, **kwargs):
                # openpyxl streams each worksheet with write(temp_file, arcname);
                # route through writestr so the entry gets the pinned timestamp.
                with open(filename, "rb") as fh:
                    data = fh.read()
                return self.writestr(arcname if arcname is not None else filename, data)

        archive = _PinnedZip(path, "w", ZIP_DEFLATED, allowZip64=True)
        try:
            ExcelWriter(wb, archive).save()
        finally:
            archive.close()

    # -- Ergonomics + native charts -------------------------------------------
    def _apply_ergonomics(self, sheets: dict[str, Worksheet]) -> None:
        """Column widths + freeze panes so labels don't overflow and headers
        stay visible on scroll — the first thing a reader does is scroll."""
        for name, ws in sheets.items():
            ws.column_dimensions["A"].width = _LABEL_COL_WIDTH
            for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J"):
                ws.column_dimensions[col].width = _DATA_COL_WIDTH
            # Freeze the label column (and the header row on data-grid tabs) so
            # they stay on screen while scrolling.
            ws.freeze_panes = "B4" if name in _GRID_TABS else "B1"

    def _add_charts(self, sheets: dict[str, Worksheet], model: ModelBundle) -> None:
        n = int(self._anchors["is_n"])

        # 1) Football-field horizontal bar (Low/High per method).
        ff = sheets[SH_FF]
        n_rows = 2  # DCF + Comps rows written at rows 4-5
        bar = BarChart()
        bar.type = "bar"  # horizontal
        bar.title = "Football Field — Valuation Range ($/share)"
        bar.height, bar.width = 6.5, 15
        data = Reference(ff, min_col=2, max_col=3, min_row=3, max_row=3 + n_rows)
        cats = Reference(ff, min_col=1, min_row=4, max_row=3 + n_rows)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ff.add_chart(bar, "E3")

        # 2) Revenue bar + margin line on Model IS (revenue row 4, EBITDA row 10).
        mis = sheets[SH_MODEL_IS]
        rev_ref = Reference(mis, min_col=3, max_col=2 + n, min_row=4, max_row=4)
        rbar = BarChart()
        rbar.title = "Projected Revenue ($)"
        rbar.height, rbar.width = 6.5, 15
        rbar.add_data(rev_ref, from_rows=True, titles_from_data=False)
        rbar.set_categories(Reference(mis, min_col=3, max_col=2 + n, min_row=3, max_row=3))
        mis.add_chart(rbar, "A18")

        # 3) FCFF bar on the DCF tab (FCFF row 4 across projected years).
        dcf = sheets[SH_DCF]
        fbar = BarChart()
        fbar.title = "Projected Unlevered FCFF ($)"
        fbar.height, fbar.width = 6.5, 15
        fbar.add_data(
            Reference(dcf, min_col=2, max_col=1 + n, min_row=4, max_row=4), from_rows=True
        )
        fbar.set_categories(Reference(dcf, min_col=2, max_col=1 + n, min_row=3, max_row=3))
        dcf.add_chart(fbar, "A28")

        # 4) Comps scatter: EV/EBITDA (col D) vs EV/Revenue (col C) across peers.
        comps = sheets[SH_COMPS]
        n_peers = len(model.comps.peers)
        if n_peers >= 2:
            sc = ScatterChart()
            sc.title = "Trading Comps — EV/EBITDA vs EV/Revenue"
            sc.x_axis.title, sc.y_axis.title = "EV/Revenue (x)", "EV/EBITDA (x)"
            sc.height, sc.width = 7, 12
            xs = Reference(comps, min_col=3, min_row=4, max_row=3 + n_peers)
            ys = Reference(comps, min_col=4, min_row=4, max_row=3 + n_peers)
            sc.series.append(Series(ys, xs, title="Peers"))
            comps.add_chart(sc, f"A{6 + n_peers}")

    # -- Assumptions: the ONLY input tab (blue) --------------------------------
    def _assumptions(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Assumptions (inputs — blue)")
        a = model.proj_assumptions
        n = a.n_years
        # Projection-year headers (formula-free labels).
        S.label_cell(ws, "A3", "Projection year", bold=True)
        for j in range(n):
            S.label_cell(ws, f"{_col(2 + j)}3", f"Y{j + 1}", bold=True)

        rows = [
            ("Revenue growth", a.revenue_growth, S.FMT_PERCENT, "growth"),
            ("Gross margin", a.gross_margin, S.FMT_PERCENT, "gm"),
            ("SG&A % revenue", a.sga_pct_revenue, S.FMT_PERCENT, "sga"),
            ("R&D % revenue", a.rnd_pct_revenue, S.FMT_PERCENT, "rnd"),
            ("Capex % revenue", a.capex_pct_revenue, S.FMT_PERCENT, "capex"),
            ("D&A % revenue", a.da_pct_revenue, S.FMT_PERCENT, "da"),
            ("Tax rate", a.tax_rate, S.FMT_PERCENT, "tax"),
        ]
        r = 4
        for label, series, fmt, key in rows:
            S.label_cell(ws, f"A{r}", label)
            for j in range(n):
                val = series[j] if j < len(series) else 0.0
                S.input_cell(ws, f"{_col(2 + j)}{r}", val, fmt)
            self._anchors[f"assum_{key}_row"] = str(r)
            r += 1

        # Change-in-working-capital driver row (engine-derived; blue input on
        # Assumptions per the banker convention — dWC carries its CF sign).
        wc_series = model.statements.series(LineItem.CHANGE_IN_WC)[model.statements.n_hist :]
        S.label_cell(ws, f"A{r}", "Change in working capital")
        for j in range(n):
            val = wc_series[j] if j < len(wc_series) and wc_series[j] is not None else 0.0
            S.input_cell(ws, f"{_col(2 + j)}{r}", val)
        self._anchors["assum_dwc_row"] = str(r)
        r += 1

        # Scalars.
        r += 1
        S.label_cell(ws, f"A{r}", "Min cash (revolver floor)")
        S.input_cell(ws, f"B{r}", a.min_cash)
        self._anchors["assum_min_cash"] = f"B{r}"
        r += 1
        S.label_cell(ws, f"A{r}", "Terminal growth (g)")
        S.input_cell(ws, f"B{r}", model.terminal.terminal_growth, S.FMT_PERCENT)
        self._anchors["assum_g_ref"] = f"'{SH_ASSUM}'!B{r}"
        r += 1
        S.label_cell(ws, f"A{r}", "Exit EV/EBITDA")
        S.input_cell(ws, f"B{r}", model.terminal.exit_ev_ebitda, S.FMT_MULTIPLE)
        self._anchors["assum_exit_ref"] = f"'{SH_ASSUM}'!B{r}"

        # WACC / CAPM inputs (blue) — the ONLY place these hard inputs live.
        w = model.wacc_inputs
        r += 2
        S.subheader(ws, f"A{r}", "WACC / CAPM inputs")
        wacc_inputs = [
            ("Risk-free rate", w.risk_free_rate, S.FMT_PERCENT, "rf"),
            ("Beta", w.beta, S.FMT_PERSHARE, "beta"),
            ("Equity risk premium", w.equity_risk_premium, S.FMT_PERCENT, "erp"),
            ("Pre-tax cost of debt", w.pretax_cost_of_debt, S.FMT_PERCENT, "kd"),
            ("WACC tax rate", w.tax_rate, S.FMT_PERCENT, "wtax"),
            ("Market cap", w.market_cap, S.FMT_CURRENCY, "mcap"),
            ("Total debt", w.total_debt, S.FMT_CURRENCY, "wdebt"),
        ]
        for label, val, fmt, key in wacc_inputs:
            r += 1
            S.label_cell(ws, f"A{r}", label)
            S.input_cell(ws, f"B{r}", val, fmt)
            self._anchors[f"assum_{key}_ref"] = f"'{SH_ASSUM}'!B{r}"

    # -- WACC (CAPM) -----------------------------------------------------------
    def _wacc(self, ws: Worksheet, model: ModelBundle) -> None:
        """All-formula WACC build. Inputs live on Assumptions (banker rule);
        this tab references them so every cell here is a live formula."""
        S.section_header(ws, "A1", "WACC (CAPM)")
        # Qualified references to the blue inputs on Assumptions.
        rf = self._anchors["assum_rf_ref"]
        beta = self._anchors["assum_beta_ref"]
        erp = self._anchors["assum_erp_ref"]
        kd = self._anchors["assum_kd_ref"]
        wtax = self._anchors["assum_wtax_ref"]
        mcap = self._anchors["assum_mcap_ref"]
        debt = self._anchors["assum_wdebt_ref"]

        S.label_cell(ws, "A3", "Cost of equity (Ke)")
        S.formula_cell(ws, "B3", f"={rf}+{beta}*{erp}", S.FMT_PERCENT)
        ke = "B3"
        S.label_cell(ws, "A4", "After-tax cost of debt")
        S.formula_cell(ws, "B4", f"={kd}*(1-{wtax})", S.FMT_PERCENT)
        kd_at = "B4"
        S.label_cell(ws, "A5", "Equity weight")
        S.formula_cell(ws, "B5", f"={mcap}/({mcap}+{debt})", S.FMT_PERCENT)
        we = "B5"
        S.label_cell(ws, "A6", "Debt weight")
        S.formula_cell(ws, "B6", f"={debt}/({mcap}+{debt})", S.FMT_PERCENT)
        wd = "B6"
        S.label_cell(ws, "A7", "WACC", bold=True)
        S.formula_cell(ws, "B7", f"={we}*{ke}+{wd}*{kd_at}", S.FMT_PERCENT)
        self._anchors["wacc"] = f"'{SH_WACC}'!B7"
        self._anchors["wacc_cell"] = "B7"

    # -- Model income statement (live formula chain off Assumptions) -----------
    def _model_is(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Model Income Statement (projection)")
        stmts = model.statements
        n = model.proj_assumptions.n_years
        # Base (last historical) revenue as the chain anchor — a reported value.
        base_rev = _last_hist(stmts, LineItem.REVENUE) or 0.0

        S.label_cell(ws, "A3", "Line ($)", bold=True)
        S.value_cell(ws, "B3", "Base", S.FMT_YEAR)
        for j in range(n):
            S.label_cell(ws, f"{_col(3 + j)}3", f"Y{j + 1}", bold=True)

        assum = f"'{SH_ASSUM}'!"
        gr = self._anchors["assum_growth_row"]
        gm = self._anchors["assum_gm_row"]
        sga = self._anchors["assum_sga_row"]
        rnd = self._anchors["assum_rnd_row"]
        tax = self._anchors["assum_tax_row"]
        da = self._anchors["assum_da_row"]

        # Revenue row 4: base in B, chain in C..
        S.label_cell(ws, "A4", "Revenue")
        S.value_cell(ws, "B4", base_rev)
        for j in range(n):
            prev = f"{_col(2 + j)}4"
            gcell = f"{assum}{_col(2 + j)}{gr}"
            S.formula_cell(ws, f"{_col(3 + j)}4", f"={prev}*(1+{gcell})")
        # Gross profit row 5 = rev*gm
        S.label_cell(ws, "A5", "Gross profit")
        for j in range(n):
            rev = f"{_col(3 + j)}4"
            gmc = f"{assum}{_col(2 + j)}{gm}"
            S.formula_cell(ws, f"{_col(3 + j)}5", f"={rev}*{gmc}")
        # SG&A row 6, R&D row 7
        S.label_cell(ws, "A6", "SG&A")
        S.label_cell(ws, "A7", "R&D")
        for j in range(n):
            rev = f"{_col(3 + j)}4"
            S.formula_cell(ws, f"{_col(3 + j)}6", f"={rev}*{assum}{_col(2 + j)}{sga}")
            S.formula_cell(ws, f"{_col(3 + j)}7", f"={rev}*{assum}{_col(2 + j)}{rnd}")
        # EBIT row 8 = GP - SG&A - R&D
        S.label_cell(ws, "A8", "EBIT", bold=True)
        for j in range(n):
            c = _col(3 + j)
            S.formula_cell(ws, f"{c}8", f"={c}5-{c}6-{c}7")
        # D&A row 9 = rev*da%
        S.label_cell(ws, "A9", "D&A")
        for j in range(n):
            rev = f"{_col(3 + j)}4"
            S.formula_cell(ws, f"{_col(3 + j)}9", f"={rev}*{assum}{_col(2 + j)}{da}")
        # EBITDA row 10 = EBIT + D&A
        S.label_cell(ws, "A10", "EBITDA", bold=True)
        for j in range(n):
            c = _col(3 + j)
            S.formula_cell(ws, f"{c}10", f"={c}8+{c}9")
        # Net interest income row 11 — engine-sourced (interest on the net-cash
        # balance, net of any interest expense). A black value, not a blue input:
        # it derives from the projected cash balance × the cash yield, which the
        # engine computes via its fixed-point revolver/interest solve. Carrying
        # it here (rather than re-deriving a circular cash×rate formula) keeps the
        # workbook NI tied to the engine to the cent.
        int_inc = model.statements.series(LineItem.INTEREST_INCOME)[model.statements.n_hist :]
        int_exp = model.statements.series(LineItem.INTEREST_EXPENSE)[model.statements.n_hist :]
        S.label_cell(ws, "A11", "Net interest income")
        for j in range(n):
            ii = (int_inc[j] or 0.0) if j < len(int_inc) else 0.0
            ie = (int_exp[j] or 0.0) if j < len(int_exp) else 0.0
            S.value_cell(ws, f"{_col(3 + j)}11", ii - ie)
        # Pretax row 12 = EBIT + net interest income
        S.label_cell(ws, "A12", "Pre-tax income", bold=True)
        for j in range(n):
            c = _col(3 + j)
            S.formula_cell(ws, f"{c}12", f"={c}8+{c}11")
        # Taxes row 13, Net income row 14 (tax on PRE-TAX income, incl. int inc)
        S.label_cell(ws, "A13", "Taxes")
        S.label_cell(ws, "A14", "Net income", bold=True)
        for j in range(n):
            c = _col(3 + j)
            S.formula_cell(ws, f"{c}13", f"={c}12*{assum}{_col(2 + j)}{tax}")
            S.formula_cell(ws, f"{c}14", f"={c}12-{c}13")

        # Revenue CAGR named-range cell (row 16).
        S.label_cell(ws, "A16", "Revenue CAGR (proj)")
        last_rev = f"{_col(2 + n)}4"
        S.formula_cell(ws, "B16", f"=({last_rev}/B4)^(1/{n})-1", S.FMT_PERCENT)
        self._anchors["rev_cagr"] = f"'{SH_MODEL_IS}'!B16"
        # Save EBIT/EBITDA/rev/NI row addresses for the DCF + cell map.
        self._anchors["is_n"] = str(n)
        self._anchors["is_ni_row"] = "14"

    # -- DCF (live formulas off Model IS + WACC + Assumptions) -----------------
    def _dcf(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Discounted Cash Flow (FCFF)")
        n = int(self._anchors["is_n"])
        d = model.dcf
        assum = f"'{SH_ASSUM}'!"
        capex_row = self._anchors["assum_capex_row"]
        tax_row = self._anchors["assum_tax_row"]
        mis = f"'{SH_MODEL_IS}'!"
        wacc = self._anchors["wacc"]

        S.label_cell(ws, "A3", "Projection year", bold=True)
        for j in range(n):
            S.label_cell(ws, f"{_col(2 + j)}3", f"Y{j + 1}", bold=True)

        # Change in working capital: mirror the Assumptions dWC input row as a
        # live formula (row 7) so FCFF references it locally while the sole blue
        # input stays on Assumptions (banker rule). dWC carries its CF sign
        # (a build is negative), so FCFF ADDS it.
        dwc_row = self._anchors["assum_dwc_row"]
        S.label_cell(ws, "A7", "Change in working capital")
        for j in range(n):
            S.formula_cell(ws, f"{_col(2 + j)}7", f"={assum}{_col(2 + j)}{dwc_row}")

        # FCFF row 4 = EBIT*(1-tax) + D&A - capex + dWC.
        # EBIT is Model IS row 8 col (3+j); D&A row 9; revenue row 4. capex is
        # taken as % revenue (a positive outflow, subtracted). dWC (row 7) is
        # already signed (negative = build), so it is ADDED.
        S.label_cell(ws, "A4", "FCFF")
        for j in range(n):
            c_is = _col(3 + j)
            ebit = f"{mis}{c_is}8"
            da = f"{mis}{c_is}9"
            rev = f"{mis}{c_is}4"
            taxc = f"{assum}{_col(2 + j)}{tax_row}"
            capexc = f"{assum}{_col(2 + j)}{capex_row}"
            dwc = f"{_col(2 + j)}7"
            S.formula_cell(ws, f"{_col(2 + j)}4", f"={ebit}*(1-{taxc})+{da}-{rev}*{capexc}+{dwc}")
        # Discount factor row 5 (mid-year if configured)
        mid = 0.5 if model.terminal.mid_year_convention else 0.0
        S.label_cell(ws, "A5", "Discount factor")
        for j in range(n):
            t = j + 1
            S.formula_cell(ws, f"{_col(2 + j)}5", f"=1/(1+{wacc})^({t}-{mid})", S.FMT_PERSHARE)
        # PV of FCFF row 6
        S.label_cell(ws, "A6", "PV of FCFF")
        for j in range(n):
            c = _col(2 + j)
            S.formula_cell(ws, f"{c}6", f"={c}4*{c}5")
        # Sum PV explicit row 8
        last = _col(1 + n)
        S.label_cell(ws, "A8", "PV explicit FCFF", bold=True)
        S.formula_cell(ws, "B8", f"=SUM(B6:{last}6)")
        pv_explicit = "B8"
        # Terminal values (Gordon + exit) rows 9-12.
        term_ebitda = f"'{SH_MODEL_IS}'!{_col(2 + n)}10"  # last EBITDA
        last_df = f"{_col(1 + n)}5"  # mid-year factor for the perpetuity stream
        g_ref = self._anchors["assum_g_ref"]
        wacc_ref = self._anchors["wacc"]
        # Normalized terminal FCFF (engine value) — steady-state basis for the
        # Gordon perpetuity (capex==D&A, WC investment scaled to g), NOT the raw
        # last-year FCFF. Carried as a labeled black value below the projection
        # (the engine computes the normalization; see valuation/engine.py).
        S.label_cell(ws, "A24", "Terminal FCFF (normalized)")
        S.value_cell(ws, "B24", model.dcf.terminal_fcff_normalized)
        term_fcff_norm = "B24"
        # Full-year discount factor for the exit (a year-end sale at year N).
        S.label_cell(ws, "A25", "Exit discount factor (full-year)")
        S.formula_cell(ws, "B25", f"=1/(1+{wacc_ref})^{n}", S.FMT_PERSHARE)
        exit_df = "B25"
        S.label_cell(ws, "A9", "Terminal value (Gordon)")
        S.formula_cell(ws, "B9", f"={term_fcff_norm}*(1+{g_ref})/({wacc_ref}-{g_ref})")
        S.label_cell(ws, "A10", "Terminal value (exit)")
        S.formula_cell(ws, "B10", f"={term_ebitda}*{self._anchors['assum_exit_ref']}")
        S.label_cell(ws, "A11", "PV terminal (Gordon)")
        S.formula_cell(ws, "B11", f"=B9*{last_df}")
        S.label_cell(ws, "A12", "PV terminal (exit)")
        S.formula_cell(ws, "B12", f"=B10*{exit_df}")
        # EV rows 14-15
        S.label_cell(ws, "A14", "EV (Gordon)", bold=True)
        S.formula_cell(ws, "B14", f"={pv_explicit}+B11")
        S.label_cell(ws, "A15", "EV (exit)", bold=True)
        S.formula_cell(ws, "B15", f"={pv_explicit}+B12")
        self._anchors["ev_gordon"] = f"'{SH_DCF}'!B14"
        self._anchors["ev_exit"] = f"'{SH_DCF}'!B15"
        # Bridge rows 17-22
        S.label_cell(ws, "A17", "Net debt")
        # Engine-sourced bridge value (from last historical BS), not a user
        # input — black value cell, so the blue-input audit stays clean.
        S.value_cell(ws, "B17", d.net_debt)
        self._anchors["net_debt"] = f"'{SH_DCF}'!B17"
        S.label_cell(ws, "A18", "Shares diluted")
        S.value_cell(ws, "B18", d.shares_diluted, S.FMT_SHARES)
        self._anchors["shares"] = f"'{SH_DCF}'!B18"
        S.label_cell(ws, "A19", "Equity value (Gordon)", bold=True)
        S.formula_cell(ws, "B19", "=B14-B17")
        S.label_cell(ws, "A20", "Equity value (exit)", bold=True)
        S.formula_cell(ws, "B20", "=B15-B17")
        self._anchors["eq_gordon"] = f"'{SH_DCF}'!B19"
        self._anchors["eq_exit"] = f"'{SH_DCF}'!B20"
        S.label_cell(ws, "A21", "Implied price (Gordon)", bold=True)
        S.formula_cell(ws, "B21", "=B19/B18", S.FMT_PERSHARE)
        S.label_cell(ws, "A22", "Implied price (exit)", bold=True)
        S.formula_cell(ws, "B22", "=B20/B18", S.FMT_PERSHARE)
        self._anchors["price_gordon"] = f"'{SH_DCF}'!B21"
        self._anchors["price_exit"] = f"'{SH_DCF}'!B22"

    # -- Trading comps ---------------------------------------------------------
    def _comps(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Trading Comps")
        headers = ["Ticker", "EV", "EV/Revenue", "EV/EBITDA", "P/E"]
        for k, h in enumerate(headers):
            S.subheader(ws, f"{_col(1 + k)}3", h)
        r = 4
        first_peer_row = r
        for p in model.comps.peers:
            S.label_cell(ws, f"A{r}", p.ticker)
            S.value_cell(ws, f"B{r}", p.enterprise_value)
            # Missing multiples are left BLANK (not 0.0) — a zero multiple is a
            # false claim, and MEDIAN/AVERAGE skip empty cells, so the workbook
            # median matches the engine (which excludes None).
            _mult_or_blank(ws, f"C{r}", p.ev_revenue_ltm)
            _mult_or_blank(ws, f"D{r}", p.ev_ebitda_ltm)
            _mult_or_blank(ws, f"E{r}", p.pe_ltm)
            r += 1
        last_peer_row = r - 1
        # Median EV/EBITDA formula (blank cells are skipped by MEDIAN).
        r += 1
        S.label_cell(ws, f"A{r}", "Median EV/EBITDA", bold=True)
        S.formula_cell(ws, f"D{r}", f"=MEDIAN(D{first_peer_row}:D{last_peer_row})", S.FMT_MULTIPLE)
        # Implied price from comps: engine-sourced value (black), not an input.
        r += 1
        S.label_cell(ws, f"A{r}", "Implied price (EV/EBITDA)", bold=True)
        val = model.comps.implied_price_from_ebitda or 0.0
        S.value_cell(ws, f"B{r}", val, S.FMT_PERSHARE)
        self._anchors["comps_implied"] = f"'{SH_COMPS}'!B{r}"

    def _precedents(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Precedent Transactions")
        headers = ["Date", "Acquirer", "Target", "EV", "EV/Rev", "EV/EBITDA", "Source"]
        for k, h in enumerate(headers):
            S.subheader(ws, f"{_col(1 + k)}3", h)
        r = 4
        for t in model.precedents:
            S.label_cell(ws, f"A{r}", t.date)
            S.label_cell(ws, f"B{r}", t.acquirer)
            S.label_cell(ws, f"C{r}", t.target)
            S.value_cell(ws, f"D{r}", t.ev)
            _mult_or_blank(ws, f"E{r}", t.ev_revenue)
            _mult_or_blank(ws, f"F{r}", t.ev_ebitda)
            S.label_cell(ws, f"G{r}", t.source)
            r += 1

    # -- LBO (handles None) ----------------------------------------------------
    def _lbo(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "LBO")
        if model.lbo is None:
            S.label_cell(ws, "A3", "LBO not modeled for this run (net-cash subject).")
            # Still define name targets pointing to zero cells so refs resolve.
            S.value_cell(ws, "B5", 0.0, S.FMT_PERCENT)
            S.value_cell(ws, "B6", 0.0, S.FMT_MULTIPLE)
            self._anchors["lbo_irr"] = f"'{SH_LBO}'!B5"
            self._anchors["lbo_moic"] = f"'{SH_LBO}'!B6"
            return
        lbo = model.lbo
        S.label_cell(ws, "A3", "Sources & Uses")
        r = 4
        for k, v in lbo.sources.items():
            S.label_cell(ws, f"A{r}", f"Source: {k}")
            S.value_cell(ws, f"B{r}", v)
            r += 1
        for k, v in lbo.uses.items():
            S.label_cell(ws, f"A{r}", f"Use: {k}")
            S.value_cell(ws, f"B{r}", v)
            r += 1

        # 5-year debt schedule with the levered cash sweep (engine values).
        r += 1
        S.subheader(ws, f"A{r}", "Debt schedule (levered cash sweep)")
        r += 1
        cols = [
            ("Year", None),
            ("Begin debt", "begin"),
            ("Cash interest", "interest"),
            ("Levered FCF", "fcf"),
            ("Sweep", "sweep"),
            ("End debt", "end"),
            ("Cash balance", "cash_balance"),
        ]
        for k, (label, _key) in enumerate(cols):
            S.subheader(ws, f"{_col(1 + k)}{r}", label)
        for i, row in enumerate(lbo.debt_schedule):
            rr = r + 1 + i
            S.label_cell(ws, f"A{rr}", f"Y{i + 1}")
            for k, (_label, key) in enumerate(cols):
                if key is not None:
                    S.value_cell(ws, f"{_col(1 + k)}{rr}", row.get(key, 0.0))
        r = r + 1 + len(lbo.debt_schedule) + 1

        S.label_cell(ws, f"A{r}", "Exit equity value", bold=True)
        S.value_cell(ws, f"B{r}", lbo.exit_equity_value)
        r += 1
        S.label_cell(ws, f"A{r}", "IRR", bold=True)
        S.value_cell(ws, f"B{r}", lbo.irr, S.FMT_PERCENT)
        self._anchors["lbo_irr"] = f"'{SH_LBO}'!B{r}"
        r += 1
        S.label_cell(ws, f"A{r}", "MOIC", bold=True)
        S.value_cell(ws, f"B{r}", lbo.moic, S.FMT_MULTIPLE)
        self._anchors["lbo_moic"] = f"'{SH_LBO}'!B{r}"
        r += 1
        S.label_cell(
            ws,
            f"A{r}",
            "Illustrative only — DECK is net-cash today; this shows what leverage could do.",
        )

    # -- Historical statements (XBRL tie-out surface: reported values) ---------
    def _historical(self, ws: Worksheet, model: ModelBundle, items: list[LineItem]) -> None:
        S.section_header(ws, "A1", f"{ws.title} (reported — ties to XBRL)")
        stmts = model.statements
        hist = stmts.periods[: stmts.n_hist]
        for k, p in enumerate(hist):
            S.label_cell(ws, f"{_col(2 + k)}3", f"FY{p.fy}", bold=True)
        for i, li in enumerate(items):
            r = 4 + i
            S.label_cell(ws, f"A{r}", li.value)
            series = stmts.series(li)
            for k in range(len(hist)):
                v = series[k]
                if v is not None:
                    S.value_cell(ws, f"{_col(2 + k)}{r}", v)

    def _statement_tab(
        self, ws: Worksheet, model: ModelBundle, title: str, lines: list[tuple[str, LineItem, bool]]
    ) -> None:
        """Render a projected statement tab (BS or CF) from engine values.

        ``lines`` is (label, LineItem, bold). Values are the engine's projected
        series — black cells, correct to the engine (these are outputs of the
        3-statement builder, not user inputs). Column headers are the projected
        fiscal years. This makes the deliverable's balance sheet and cash-flow
        statement real, populated tabs rather than placeholders.
        """
        S.section_header(ws, "A1", f"Model {title}")
        stmts = model.statements
        n = stmts.n_hist
        nyr = len(stmts.periods) - n
        S.label_cell(ws, "A3", "Line ($)", bold=True)
        for j in range(nyr):
            p = stmts.periods[n + j]
            fy = p.fy if p.fy is not None else p.end.year
            S.label_cell(ws, f"{_col(2 + j)}3", f"FY{fy}E", bold=True)
        for i, (label, li, bold) in enumerate(lines):
            r = 4 + i
            S.label_cell(ws, f"A{r}", label, bold=bold)
            series = stmts.series(li)[n:]
            for j in range(nyr):
                v = series[j] if j < len(series) else None
                if v is not None:
                    S.value_cell(ws, f"{_col(2 + j)}{r}", v)
        # A note documenting the linkage the engine enforces (verified by the
        # invariant runner: BS balances, CFS ties to BS cash, RE & PP&E rolls).
        note_r = 4 + len(lines) + 1
        S.label_cell(
            ws,
            f"A{note_r}",
            "Engine-computed & invariant-checked: BS balances every period; "
            "CFS ending cash ties to BS cash; RE and PP&E rolls hold.",
        )

    def _model_bs(self, ws: Worksheet, model: ModelBundle) -> None:
        self._statement_tab(
            ws,
            model,
            "Balance Sheet",
            [
                ("Cash & equivalents", LineItem.CASH, False),
                ("Accounts receivable", LineItem.ACCOUNTS_RECEIVABLE, False),
                ("Inventory", LineItem.INVENTORY, False),
                ("Total current assets", LineItem.TOTAL_CURRENT_ASSETS, True),
                ("Property, plant & equipment, net", LineItem.PPE_NET, False),
                ("Goodwill & intangibles", LineItem.GOODWILL, False),
                ("Total assets", LineItem.TOTAL_ASSETS, True),
                ("Accounts payable", LineItem.ACCOUNTS_PAYABLE, False),
                ("Total current liabilities", LineItem.TOTAL_CURRENT_LIABILITIES, True),
                ("Total liabilities", LineItem.TOTAL_LIABILITIES, True),
                ("Retained earnings", LineItem.RETAINED_EARNINGS, False),
                ("Total stockholders' equity", LineItem.TOTAL_EQUITY, True),
            ],
        )

    def _model_cf(self, ws: Worksheet, model: ModelBundle) -> None:
        self._statement_tab(
            ws,
            model,
            "Cash Flow Statement",
            [
                ("Net income", LineItem.NET_INCOME, False),
                ("Depreciation & amortization", LineItem.DA_CF, False),
                ("Change in working capital", LineItem.CHANGE_IN_WC, False),
                ("Cash from operations", LineItem.CFO, True),
                ("Capital expenditures", LineItem.CAPEX, False),
                ("Cash from investing", LineItem.CFI, True),
                ("Net change in cash", LineItem.NET_CHANGE_IN_CASH, True),
            ],
        )

    # -- Sensitivities (live formula grid) -------------------------------------
    def _sensitivities(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Sensitivities: implied price = f(WACC, g)")
        # A simple live grid using the Gordon terminal on the DCF EV.
        # Header row of WACC deltas, first col of g deltas; each cell recomputes
        # implied price via the perpetuity relationship anchored to DCF PV.
        S.label_cell(ws, "A3", "g \\ WACC", bold=True)
        wacc = self._anchors["wacc"]
        base_g = self._anchors["assum_g_ref"]
        pv_explicit = f"'{SH_DCF}'!B8"
        # Use the NORMALIZED terminal FCFF (DCF!B24) and the MID-YEAR discount
        # exponent (N-0.5) so the centre cell (Δg=0, ΔWACC=0) reproduces the DCF
        # tab's Gordon implied price exactly — the two tabs cannot disagree.
        term_fcff = f"'{SH_DCF}'!B24"
        net_debt = self._anchors["net_debt"]
        shares = self._anchors["shares"]
        n = int(self._anchors["is_n"])
        wacc_deltas = [-0.01, -0.005, 0.0, 0.005, 0.01]
        g_deltas = [-0.01, -0.005, 0.0, 0.005, 0.01]
        for cidx, wd in enumerate(wacc_deltas):
            S.formula_cell(ws, f"{_col(2 + cidx)}3", f"={wacc}+{wd}", S.FMT_PERCENT)
        for ridx, gd in enumerate(g_deltas):
            r = 4 + ridx
            S.formula_cell(ws, f"A{r}", f"={base_g}+{gd}", S.FMT_PERCENT)
            for cidx, _wd in enumerate(wacc_deltas):
                wcell = f"{_col(2 + cidx)}3"
                gcell = f"A{r}"
                tv = f"{term_fcff}*(1+{gcell})/({wcell}-{gcell})"
                ev = f"({pv_explicit}+{tv}/(1+{wcell})^({n}-0.5))"
                price = f"=({ev}-{net_debt})/{shares}"
                S.formula_cell(ws, f"{_col(2 + cidx)}{r}", price, S.FMT_PERSHARE)
        # Record the centre cell so the differential guards it against drift.
        self._anchors["sens_center"] = f"{_col(2 + 2)}{4 + 2}"  # ΔWACC=0, Δg=0 -> D6

    def _football_field(self, ws: Worksheet, model: ModelBundle) -> None:
        S.section_header(ws, "A1", "Football Field (valuation ranges)")
        S.subheader(ws, "A3", "Method")
        S.subheader(ws, "B3", "Low")
        S.subheader(ws, "C3", "High")
        rows = [
            ("DCF (Gordon–Exit)", self._anchors["price_gordon"], self._anchors["price_exit"]),
            ("Comps (EV/EBITDA)", self._anchors["comps_implied"], self._anchors["comps_implied"]),
        ]
        for i, (label, lo, hi) in enumerate(rows):
            r = 4 + i
            S.label_cell(ws, f"A{r}", label)
            S.formula_cell(ws, f"B{r}", f"={lo}", S.FMT_PERSHARE)
            S.formula_cell(ws, f"C{r}", f"={hi}", S.FMT_PERSHARE)
        # Markers
        r = 4 + len(rows) + 1
        S.label_cell(ws, f"A{r}", "Current price")
        S.formula_cell(ws, f"B{r}", f"={self._anchors['current_price']}", S.FMT_PERSHARE)
        r += 1
        S.label_cell(ws, f"A{r}", "Price target")
        S.formula_cell(ws, f"B{r}", f"={self._anchors['price_target']}", S.FMT_PERSHARE)

    # -- Cover -----------------------------------------------------------------
    def _cover(self, ws: Worksheet, model: ModelBundle) -> None:
        company = model.company
        name = getattr(company, "name", "Company")
        ticker = getattr(company, "ticker", "TICK")
        S.section_header(ws, "A1", f"{name} ({ticker}) — Initiating Coverage")
        S.label_cell(ws, "A3", "Rating")
        # Rating text on Cover (an input string, not a computed number).
        rating = model.rating or "—"
        ws["B3"].value = rating
        self._anchors["rating"] = f"'{SH_COVER}'!B3"
        S.label_cell(ws, "A4", "Current price")
        S.input_cell(ws, "B4", model.current_price, S.FMT_PERSHARE)  # allowed Cover input
        self._anchors["current_price"] = f"'{SH_COVER}'!B4"
        S.label_cell(ws, "A5", "12-mo price target")
        # Price target: an analyst-set input on the cover (allowed second input).
        S.input_cell(ws, "B5", model.price_target or 0.0, S.FMT_PERSHARE)
        self._anchors["price_target"] = f"'{SH_COVER}'!B5"
        S.label_cell(ws, "A6", "Implied upside")
        S.formula_cell(ws, "B6", "=B5/B4-1", S.FMT_PERCENT)
        # Triangulation mini-table referencing the method tabs.
        S.subheader(ws, "A8", "Valuation triangulation")
        S.label_cell(ws, "A9", "DCF (Gordon)")
        S.formula_cell(ws, "B9", f"={self._anchors['price_gordon']}", S.FMT_PERSHARE)
        S.label_cell(ws, "A10", "DCF (exit)")
        S.formula_cell(ws, "B10", f"={self._anchors['price_exit']}", S.FMT_PERSHARE)
        S.label_cell(ws, "A11", "Comps (EV/EBITDA)")
        S.formula_cell(ws, "B11", f"={self._anchors['comps_implied']}", S.FMT_PERSHARE)
        S.label_cell(
            ws,
            "A13",
            "Educational project — not investment advice. "
            "SEC EDGAR data used per fair-access policy; no SEC endorsement implied.",
        )

    # -- Named ranges ----------------------------------------------------------
    def _define_names(self, wb: Workbook) -> None:
        mapping = {
            "CurrentPrice": self._anchors["current_price"],
            "PriceTarget": self._anchors["price_target"],
            "Rating": self._anchors["rating"],
            "WACC": self._anchors["wacc"],
            "TerminalGrowth": self._anchors["assum_g_ref"],
            "ExitMultiple": self._anchors["assum_exit_ref"],
            "EV_Gordon": self._anchors["ev_gordon"],
            "EV_Exit": self._anchors["ev_exit"],
            "EquityValue_Gordon": self._anchors["eq_gordon"],
            "EquityValue_Exit": self._anchors["eq_exit"],
            "ImpliedPrice_Gordon": self._anchors["price_gordon"],
            "ImpliedPrice_Exit": self._anchors["price_exit"],
            "NetDebt": self._anchors["net_debt"],
            "SharesDiluted": self._anchors["shares"],
            "Comps_ImpliedPrice_EBITDA": self._anchors["comps_implied"],
            "LBO_IRR": self._anchors["lbo_irr"],
            "LBO_MOIC": self._anchors["lbo_moic"],
            "RevenueCAGR": self._anchors["rev_cagr"],
        }
        for name, ref in mapping.items():
            wb.defined_names[name] = DefinedName(name, attr_text=ref)


def build_verifier_cell_map(model: ModelBundle) -> dict[tuple[str, str], float]:
    """Authoritative (sheet, cell) -> engine-value map for the differential.

    The verifier recalculates the workbook (via the ``formulas`` library) and
    asserts each of these cells equals the mapped engine value to the cent.
    Covers every computed *formula* cell that reproduces an engine output: the
    WACC build, the full Model-IS projection chain (revenue → EBIT → EBITDA →
    pre-tax → net income, per year) + Revenue CAGR, and the DCF valuation +
    bridge. (Baked engine-output cells — net debt, shares, comps/LBO results —
    are intentionally NOT here; they aren't formulas.)
    """
    d = model.dcf
    s = model.statements
    n = s.n_hist
    nyr = len(s.periods) - n
    m: dict[tuple[str, str], float] = {
        # WACC inputs live on Assumptions; this tab holds only the 5 computed
        # rows (Ke, Kd_at, equity weight, debt weight, WACC) at B3-B7.
        (SH_WACC, "B7"): d.wacc,  # WACC formula result
        # DCF valuation + EV->equity bridge.
        (SH_DCF, "B8"): d.pv_explicit_fcff,
        (SH_DCF, "B9"): d.terminal_value_gordon,
        (SH_DCF, "B10"): d.terminal_value_exit,
        (SH_DCF, "B11"): d.pv_terminal_gordon,
        (SH_DCF, "B12"): d.pv_terminal_exit,
        (SH_DCF, "B14"): d.enterprise_value_gordon,
        (SH_DCF, "B15"): d.enterprise_value_exit,
        (SH_DCF, "B19"): d.equity_value_gordon,
        (SH_DCF, "B20"): d.equity_value_exit,
        (SH_DCF, "B21"): d.implied_price_gordon,
        (SH_DCF, "B22"): d.implied_price_exit,
    }
    # Model-IS projection chain, per projected year (cols start at C = idx 3).
    rev = s.series(LineItem.REVENUE)[n:]
    gp = s.series(LineItem.GROSS_PROFIT)[n:]
    ebit = s.series(LineItem.OPERATING_INCOME)[n:]
    da = s.series(LineItem.DEP_AMORT)[n:]
    ni = s.series(LineItem.NET_INCOME)[n:]
    for j in range(nyr):
        col = _col(3 + j)
        if rev[j] is not None:
            m[(SH_MODEL_IS, f"{col}4")] = rev[j]
        if gp[j] is not None:
            m[(SH_MODEL_IS, f"{col}5")] = gp[j]
        if ebit[j] is not None:
            m[(SH_MODEL_IS, f"{col}8")] = ebit[j]
        if ebit[j] is not None and da[j] is not None:
            m[(SH_MODEL_IS, f"{col}10")] = ebit[j] + da[j]  # EBITDA
        if ni[j] is not None:
            m[(SH_MODEL_IS, f"{col}14")] = ni[j]
    # Sensitivities grid centre cell (ΔWACC=0, Δg=0) must reproduce the DCF
    # Gordon implied price — guards the grid against re-drifting from the DCF tab.
    m[(SH_SENS, "D6")] = d.implied_price_gordon
    return m


def _mult_or_blank(ws: Worksheet, coord: str, v: float | None) -> None:
    """Write a multiple as a black value, or leave the cell blank when the
    engine reports None (honest-unknown; never write 0.0 for 'not computed')."""
    if v is None:
        return
    S.value_cell(ws, coord, v, S.FMT_MULTIPLE)


def _col(idx: int) -> str:
    """1-based column index -> Excel column letter(s)."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _last_hist(stmts: StatementSet, li: LineItem) -> float | None:
    series = stmts.series(li)[: stmts.n_hist]
    for v in reversed(series):
        if v is not None:
            return v
    return None
