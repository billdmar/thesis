"""Excel banker-convention audit gate (G2).

Opens the workbook with openpyxl (``data_only=False`` so formulas are visible as
formulas) and enforces the house rules from docs/WORKBOOK_SPEC.md:

* **No baked values in formula regions.** Every cell the differential maps
  (:func:`~src.workbook.build_verifier_cell_map`) must be a live formula
  (``data_type == 'f'``) — a hard-coded number where a formula is required is a
  defect. This sub-check needs the model to build the map; skipped when no model
  is supplied.
* **Blue inputs are confined.** A blue-font cell (input) may appear ONLY on the
  Assumptions tab and the two allowed Cover market inputs (``B4``/``B5``).
* **Named ranges present and unambiguous.** Every name in
  :data:`~src.workbook.writer.NAMED_RANGES` exists and resolves to exactly one
  cell.

Returns an :class:`AuditReport` with ``passed`` and a list of violation strings.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import load_workbook

from src.interfaces import ModelBundle
from src.workbook import build_verifier_cell_map
from src.workbook.styles import BLUE
from src.workbook.writer import NAMED_RANGES, SH_ASSUM, SH_COVER

# The two market-data inputs allowed off the Assumptions tab (Cover current
# price + price target), per WORKBOOK_SPEC §Banker conventions.
_ALLOWED_COVER_INPUTS = {"B4", "B5"}


@dataclass
class AuditReport:
    """Result of :func:`audit_workbook`. ``passed`` iff no violations."""

    violations: list[str] = field(default_factory=list)
    cells_checked: int = 0

    @property
    def passed(self) -> bool:
        return not self.violations

    def summary(self) -> str:
        head = f"Excel audit: {self.cells_checked} checks"
        return head + (" — PASS" if self.passed else f" — {len(self.violations)} VIOLATION")


def _is_blue(cell) -> bool:
    """True when the cell's font color is the banker blue (input marker)."""
    font = cell.font
    if font is None or font.color is None:
        return False
    rgb = font.color.rgb
    # openpyxl stores an 8-digit ARGB (e.g. "000000CC") or the theme/raw value;
    # match on the trailing 6 hex digits so an alpha prefix does not hide it.
    return isinstance(rgb, str) and rgb.upper().endswith(BLUE)


def audit_workbook(workbook_path: str, model: ModelBundle | None = None) -> AuditReport:
    """Audit ``workbook_path`` for banker-convention compliance.

    When ``model`` is provided, also asserts every differential-mapped cell is a
    live formula; otherwise that sub-check is skipped (blue-font + named-range
    checks still run).
    """
    wb = load_workbook(workbook_path, data_only=False)
    violations: list[str] = []
    checks = 0

    # (a) Formula regions must be live formulas, not baked numbers.
    if model is not None:
        for (sheet, coord), _engine_value in build_verifier_cell_map(model).items():
            checks += 1
            cell = wb[sheet][coord]
            if cell.data_type != "f":
                violations.append(
                    f"{sheet}!{coord} must be a live formula but is "
                    f"data_type={cell.data_type!r} (baked value)"
                )

    # (b) Blue inputs only on Assumptions + the two allowed Cover cells.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not _is_blue(cell):
                    continue
                checks += 1
                if ws.title == SH_ASSUM:
                    continue
                if ws.title == SH_COVER and cell.coordinate in _ALLOWED_COVER_INPUTS:
                    continue
                violations.append(
                    f"blue input {ws.title}!{cell.coordinate} outside the Assumptions tab "
                    f"(only {SH_COVER}!{'/'.join(sorted(_ALLOWED_COVER_INPUTS))} allowed elsewhere)"
                )

    # (c) Every required named range exists and resolves to exactly one cell.
    for name in NAMED_RANGES:
        checks += 1
        dn = wb.defined_names.get(name)
        if dn is None:
            violations.append(f"missing named range {name!r}")
            continue
        dests = list(dn.destinations)
        if len(dests) != 1:
            violations.append(
                f"named range {name!r} resolves to {len(dests)} destinations, expected 1"
            )
            continue
        _sheet, coord = dests[0]
        if ":" in coord:
            violations.append(f"named range {name!r} spans a range {coord!r}, expected one cell")

    return AuditReport(violations=violations, cells_checked=checks)
