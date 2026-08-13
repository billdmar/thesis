"""Tests for the Excel banker-convention audit (src/verify/audit.py) and the
report-number lint (src/verify/report_lint.py).

The audit is run against a real DECK workbook (built from committed fixtures)
with and without the model, and against a deliberately-corrupted copy that
injects a blue input off the Assumptions tab.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Font
from src.comps import CompsEngine
from src.edgar import load_normalized_facts
from src.interfaces import (
    ModelBundle,
    ProjectionAssumptions,
    TerminalAssumptions,
    WACCInputs,
)
from src.statements import ThreeStatementBuilder
from src.valuation import DCFValuationEngine
from src.verify.audit import audit_workbook
from src.verify.report_lint import collect_engine_numbers, lint_report_numbers
from src.workbook import ExcelWorkbookWriter
from src.workbook.styles import BLUE


def _build_bundle() -> ModelBundle:
    nf = load_normalized_facts("DECK")
    hist = ThreeStatementBuilder().build_historical(nf)
    a = ProjectionAssumptions(
        n_years=5,
        revenue_growth=[0.10, 0.08, 0.07, 0.06, 0.05],
        gross_margin=[0.577] * 5,
        sga_pct_revenue=[0.345] * 5,
        rnd_pct_revenue=[0.0] * 5,
        dso=[40] * 5,
        dio=[110] * 5,
        dpo=[45] * 5,
        capex_pct_revenue=[0.016] * 5,
        da_pct_revenue=[0.018] * 5,
        tax_rate=[0.235] * 5,
        interest_rate_on_debt=0.0,
        interest_rate_on_cash=0.04,
        min_cash=400e6,
        dividend_payout=[0.0] * 5,
    )
    proj = ThreeStatementBuilder().project(hist, a)
    w = WACCInputs(
        risk_free_rate=0.043,
        beta=1.05,
        equity_risk_premium=0.05,
        pretax_cost_of_debt=0.06,
        tax_rate=0.235,
        market_cap=18e9,
        total_debt=0.0,
    )
    t = TerminalAssumptions(terminal_growth=0.03, exit_ev_ebitda=16.0, mid_year_convention=True)
    dcf = DCFValuationEngine().dcf(proj, w, t)
    comps = CompsEngine().build_peer_multiples(nf, [], market_data=None)
    return ModelBundle(
        company=nf.company,
        statements=proj,
        proj_assumptions=a,
        wacc_inputs=w,
        terminal=t,
        dcf=dcf,
        comps=comps,
        precedents=[],
        lbo=None,
        current_price=100.0,
        price_target=None,
        rating=None,
    )


def _write(tmp_path) -> str:
    bundle = _build_bundle()
    path = tmp_path / "DECK_model.xlsx"
    ExcelWorkbookWriter().write(str(path), bundle)
    return str(path)


def test_audit_passes_with_model_on_compliant_workbook(tmp_path):
    # The writer now emits blue inputs ONLY on Assumptions + Cover B4/B5, so its
    # unmodified output passes the audit (this locks the maintainer's fix — a regression
    # that put a blue input back on WACC/DCF/Comps would fail here).
    path = _write(tmp_path)
    report = audit_workbook(path, model=_build_bundle())
    assert report.passed, report.violations
    # Formula-region check (12 cells) + blue-font + 18 named ranges all ran.
    assert report.cells_checked > 12


def test_audit_passes_without_model_on_compliant_workbook(tmp_path):
    path = _write(tmp_path)
    report = audit_workbook(path)  # skips the formula-region sub-check
    assert report.passed, report.violations


def test_audit_flags_blue_input_off_assumptions(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    # Inject a blue (input) number onto the DCF tab — a banker-convention breach.
    cell = wb["DCF"]["E30"]
    cell.value = 123.0
    cell.font = Font(color=BLUE)
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("blue input" in v and "DCF!E30" in v for v in report.violations)


def test_audit_flags_missing_named_range(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    del wb.defined_names["WACC"]
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("missing named range 'WACC'" in v for v in report.violations)


def test_audit_flags_baked_value_in_formula_region(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    # Overwrite a mapped formula cell with a baked number (formula -> value).
    wb["DCF"]["B8"] = 42.0
    wb.save(path)

    report = audit_workbook(path, model=_build_bundle())
    assert not report.passed
    assert any("DCF!B8" in v and "baked value" in v for v in report.violations)


# --- report-number lint ---------------------------------------------------


def test_lint_passes_when_all_numbers_sourced():
    engine = {100.0, 250.5, 42.0}
    rendered = {100.0, 42.0}
    report = lint_report_numbers(rendered, engine)
    assert report.passed
    assert report.numbers_checked == 2


def test_lint_flags_unsourced_number():
    engine = {100.0, 250.5}
    rendered = {100.0, 999.99}  # 999.99 has no engine source
    report = lint_report_numbers(rendered, engine)
    assert not report.passed
    assert 999.99 in report.unsourced


def test_lint_tolerance_allows_near_matches():
    engine = {100.0}
    rendered = {100.005}  # within default abs_tol
    assert lint_report_numbers(rendered, engine).passed
    assert not lint_report_numbers(rendered, engine, abs_tol=1e-6).passed


def test_report_summaries_reflect_pass_and_fail():
    from src.verify.audit import AuditReport
    from src.verify.report_lint import LintReport

    assert "PASS" in AuditReport(cells_checked=3).summary()
    assert "VIOLATION" in AuditReport(violations=["x"], cells_checked=3).summary()
    assert "PASS" in LintReport(numbers_checked=2).summary()
    assert "UNSOURCED" in LintReport(numbers_checked=2, unsourced=[9.0]).summary()


def test_audit_flags_named_range_spanning_a_range(tmp_path):
    from openpyxl.workbook.defined_name import DefinedName

    path = _write(tmp_path)
    wb = load_workbook(path)
    # Repoint a required name at a multi-cell range — the audit wants one cell.
    wb.defined_names["WACC"] = DefinedName("WACC", attr_text="'WACC'!$B$3:$B$4")
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("spans a range" in v for v in report.violations)


def test_collect_engine_numbers_feeds_the_lint(tmp_path):
    bundle = _build_bundle()
    engine_numbers = collect_engine_numbers(bundle)
    # The DCF implied prices and EVs are all present -> a report citing them lints clean.
    rendered = {
        bundle.dcf.implied_price_gordon,
        bundle.dcf.enterprise_value_gordon,
        bundle.current_price,
    }
    report = lint_report_numbers(rendered, engine_numbers)
    assert report.passed
    assert len(engine_numbers) > 5
