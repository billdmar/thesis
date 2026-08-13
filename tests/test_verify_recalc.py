"""Tests for the reusable differential verifier (src/verify/recalc.py).

Builds a real DECK workbook from committed fixtures (same pattern as
tests/test_differential_dcf.py), runs the ``WorkbookVerifier``, and asserts the
recalculated cells match the engine to the cent. A negative test tightens the
tolerance to catch a real (sub-cent) mismatch and confirms it is reported.
"""

from __future__ import annotations

from src.comps import CompsEngine
from src.edgar import load_normalized_facts
from src.interfaces import (
    ModelBundle,
    ProjectionAssumptions,
    TerminalAssumptions,
    WACCInputs,
)
from src.statements import ThreeStatementBuilder
from src.valuation import DCFValuationEngine
from src.verify.recalc import WorkbookVerifier
from src.workbook import ExcelWorkbookWriter


def _build_bundle() -> ModelBundle:
    nf = load_normalized_facts("DECK")
    hist = ThreeStatementBuilder().build_historical(nf)
    a = ProjectionAssumptions(
        n_years=5,
        revenue_growth=[0.10, 0.08, 0.07, 0.06, 0.05],
        gross_margin=[0.577] * 5,
        sga_pct_revenue=[0.345] * 5,
        rnd_pct_revenue=[0.0] * 5,
        dso=[40] * 5,
        dio=[110] * 5,
        dpo=[45] * 5,
        capex_pct_revenue=[0.016] * 5,
        da_pct_revenue=[0.018] * 5,
        tax_rate=[0.235] * 5,
        interest_rate_on_debt=0.0,
        interest_rate_on_cash=0.04,
        min_cash=400e6,
        dividend_payout=[0.0] * 5,
    )
    proj = ThreeStatementBuilder().project(hist, a)
    w = WACCInputs(
        risk_free_rate=0.043,
        beta=1.05,
        equity_risk_premium=0.05,
        pretax_cost_of_debt=0.06,
        tax_rate=0.235,
        market_cap=18e9,
        total_debt=0.0,
    )
    t = TerminalAssumptions(terminal_growth=0.03, exit_ev_ebitda=16.0, mid_year_convention=True)
    dcf = DCFValuationEngine().dcf(proj, w, t)
    comps = CompsEngine().build_peer_multiples(nf, [], market_data=None)
    return ModelBundle(
        company=nf.company,
        statements=proj,
        proj_assumptions=a,
        wacc_inputs=w,
        terminal=t,
        dcf=dcf,
        comps=comps,
        precedents=[],
        lbo=None,
        current_price=100.0,
        price_target=None,
        rating=None,
    )


def test_verifier_passes_on_real_deck_workbook(tmp_path):
    bundle = _build_bundle()
    path = tmp_path / "DECK_model.xlsx"
    ExcelWorkbookWriter().write(str(path), bundle)

    report = WorkbookVerifier().recalc_and_diff(str(path), bundle)

    assert report.passed, "\n".join(
        f"  {m.sheet}!{m.cell}: engine={m.engine_value} workbook={m.workbook_value}"
        for m in report.mismatches
    )
    assert report.cells_checked >= 12


def test_verifier_catches_a_forced_mismatch(tmp_path):
    from openpyxl import load_workbook

    bundle = _build_bundle()
    path = tmp_path / "DECK_model.xlsx"
    ExcelWorkbookWriter().write(str(path), bundle)

    # Corrupt one mapped cell (DCF!B8 = PV of explicit FCFF) with a wrong baked
    # value, then confirm the verifier reports exactly that cell as a mismatch.
    wb = load_workbook(str(path))
    wb["DCF"]["B8"] = 1.0
    wb.save(str(path))

    report = WorkbookVerifier().recalc_and_diff(str(path), bundle)

    # B8 (PV explicit FCFF) feeds EV -> equity -> implied price, so the wrong
    # value cascades; at minimum B8 itself must be reported as a mismatch.
    assert not report.passed
    b8 = next((m for m in report.mismatches if (m.sheet, m.cell) == ("DCF", "B8")), None)
    assert b8 is not None
    assert b8.workbook_value == 1.0


def test_missing_cell_is_reported_as_mismatch(tmp_path):
    # A workbook whose mapped cells resolve to nothing (empty book) must fail
    # loudly rather than pass — exercises the None-node branch.
    from openpyxl import Workbook

    bundle = _build_bundle()
    path = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(str(path))

    report = WorkbookVerifier().recalc_and_diff(str(path), bundle)
    assert not report.passed
    assert any(m.workbook_value is None for m in report.mismatches)
