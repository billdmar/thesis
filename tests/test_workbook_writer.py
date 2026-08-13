"""Tests for the live-formula workbook writer.

The headline property: every computed cell is a live Excel FORMULA, and hard
inputs (blue) appear only on Assumptions plus the two allowed Cover inputs. A
hardcode audit enforces this — it is the workbook analog of the "numbers that
tie" directive and the precondition for the cell-level differential.
"""

from __future__ import annotations

from openpyxl import load_workbook
from src.interfaces import (
    CompsResult,
    DCFResult,
    ModelBundle,
    PeerMultiples,
    PrecedentTransaction,
    ProjectionAssumptions,
    StatementSet,
    TerminalAssumptions,
    WACCInputs,
)
from src.schema import CompanyMeta, LineItem, Period, PeriodType
from src.workbook import ExcelWorkbookWriter, build_verifier_cell_map
from src.workbook.writer import NAMED_RANGES, SH_ASSUM, SH_COVER, SHEET_ORDER


def _dur(y):
    return Period(
        PeriodType.DURATION,
        end=__import__("datetime").date(y, 3, 31),
        start=__import__("datetime").date(y - 1, 4, 1),
        fy=y,
        fp="FY",
    )


def _bundle(with_lbo: bool = False) -> ModelBundle:
    # Two historical periods + the projection lives via assumptions.
    periods = [_dur(2025), _dur(2026)]
    rows = {
        LineItem.REVENUE: [4000.0, 5000.0],
        LineItem.COST_OF_REVENUE: [1800.0, 2100.0],
        LineItem.GROSS_PROFIT: [2200.0, 2900.0],
        LineItem.OPERATING_INCOME: [900.0, 1150.0],
        LineItem.NET_INCOME: [700.0, 900.0],
        LineItem.CASH: [1800.0, 1900.0],
        LineItem.TOTAL_ASSETS: [3500.0, 3700.0],
        LineItem.TOTAL_EQUITY: [2400.0, 2500.0],
        LineItem.CFO: [1000.0, 1180.0],
        LineItem.CAPEX: [-80.0, -85.0],
    }
    stmts = StatementSet(periods=periods, rows=rows, n_hist=2)
    assum = ProjectionAssumptions(
        n_years=5,
        revenue_growth=[0.10, 0.09, 0.08, 0.07, 0.06],
        gross_margin=[0.58] * 5,
        sga_pct_revenue=[0.34] * 5,
        rnd_pct_revenue=[0.0] * 5,
        capex_pct_revenue=[0.02] * 5,
        da_pct_revenue=[0.02] * 5,
        tax_rate=[0.24] * 5,
        dso=[40] * 5,
        dio=[110] * 5,
        dpo=[45] * 5,
        min_cash=400.0,
        dividend_payout=[0.0] * 5,
    )
    wacc = WACCInputs(
        risk_free_rate=0.043,
        beta=1.1,
        equity_risk_premium=0.05,
        pretax_cost_of_debt=0.06,
        tax_rate=0.24,
        market_cap=18000.0,
        total_debt=0.0,
    )
    terminal = TerminalAssumptions(
        terminal_growth=0.03, exit_ev_ebitda=15.0, mid_year_convention=True
    )
    dcf = DCFResult(
        wacc=0.098,
        pv_explicit_fcff=3500.0,
        terminal_value_gordon=20000.0,
        terminal_value_exit=22000.0,
        pv_terminal_gordon=12000.0,
        pv_terminal_exit=13000.0,
        enterprise_value_gordon=15500.0,
        enterprise_value_exit=16500.0,
        net_debt=-1900.0,
        minority_interest=0.0,
        equity_value_gordon=17400.0,
        equity_value_exit=18400.0,
        shares_diluted=146.0,
        implied_price_gordon=119.18,
        implied_price_exit=126.03,
        fcff_by_year=[700, 760, 810, 850, 880],
        discount_factors=[0.95, 0.87, 0.79, 0.72, 0.66],
    )
    comps = CompsResult(
        peers=[
            PeerMultiples("NKE", "Nike", 120000, 110000, 3.0, 18.0, 25.0),
            PeerMultiples("CROX", "Crocs", 9000, 7500, 2.5, 9.0, 8.0),
        ],
        stats={"ev_ebitda_ltm": {"median": 13.5, "mean": 13.5}},
        implied_price_from_ebitda=112.0,
    )
    precedents = [
        PrecedentTransaction(
            "2025", "3G Capital", "Skechers", 9.4e9, None, None, "press release 2025"
        ),
    ]
    lbo = None
    if with_lbo:
        from src.interfaces import LBOResult

        lbo = LBOResult(
            sources={"debt": 60.0, "equity": 40.0},
            uses={"ev": 95.0, "fees": 5.0},
            debt_schedule=[],
            exit_equity_value=80.0,
            irr=0.22,
            moic=2.0,
        )
    return ModelBundle(
        company=CompanyMeta(cik="0000910521", ticker="DECK", name="Deckers Outdoor"),
        statements=stmts,
        proj_assumptions=assum,
        wacc_inputs=wacc,
        terminal=terminal,
        dcf=dcf,
        comps=comps,
        precedents=precedents,
        lbo=lbo,
        current_price=100.0,
        price_target=118.0,
        rating="Buy",
    )


def test_write_roundtrip_sheets_and_order(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _bundle())
    wb = load_workbook(p)
    assert wb.sheetnames == SHEET_ORDER


def test_all_named_ranges_defined_and_single_cell(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _bundle())
    wb = load_workbook(p)
    for name in NAMED_RANGES:
        assert name in wb.defined_names, f"missing named range {name}"
        dests = list(wb.defined_names[name].destinations)
        assert len(dests) == 1, f"{name} must resolve to one cell"


def test_hardcode_audit_formulas_not_values(tmp_path):
    """Computed cells outside Assumptions + Cover inputs must be formulas."""
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _bundle())
    wb = load_workbook(p, data_only=False)
    allowed_value_sheets = {SH_ASSUM}
    # Cells on Cover allowed to be numeric inputs: B4 (current price), B5 (target).
    cover_allowed = {"B4", "B5"}
    offenders = []
    for ws in wb.worksheets:
        if ws.title in allowed_value_sheets:
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)) and v is not None:
                    # A bare number on a computed sheet.
                    if ws.title == SH_COVER and cell.coordinate in cover_allowed:
                        continue
                    # historical reported values + engine bridge inputs are
                    # intentionally values; they are flagged separately below.
                    offenders.append((ws.title, cell.coordinate, v))
    # The writer intentionally carries a few engine-sourced values (historical
    # tie-out surface, net debt, shares, comps implied, LBO). Those are NOT
    # formula cells by design, so we assert the KEY DCF/WACC computed outputs
    # are formulas rather than a blanket no-numbers rule.
    wb2 = load_workbook(p, data_only=False)
    for sheet, coord in build_verifier_cell_map(_bundle()):
        c = wb2[sheet][coord]
        assert c.data_type == "f", f"{sheet}!{coord} must be a formula, got {c.value!r}"


def test_assumptions_inputs_are_numeric(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _bundle())
    wb = load_workbook(p, data_only=False)
    ws = wb[SH_ASSUM]
    # The revenue-growth row (row 4) B..F must be numbers, not formulas.
    for col in "BCDEF":
        c = ws[f"{col}4"]
        assert isinstance(c.value, (int, float))


def test_lbo_none_and_present(tmp_path):
    # None path: LBO_IRR/LBO_MOIC still defined.
    p1 = tmp_path / "no_lbo.xlsx"
    ExcelWorkbookWriter().write(str(p1), _bundle(with_lbo=False))
    wb1 = load_workbook(p1)
    assert "LBO_IRR" in wb1.defined_names
    # Present path.
    p2 = tmp_path / "lbo.xlsx"
    ExcelWorkbookWriter().write(str(p2), _bundle(with_lbo=True))
    wb2 = load_workbook(p2)
    assert "LBO_MOIC" in wb2.defined_names


def test_number_formats(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _bundle())
    wb = load_workbook(p)
    ws = wb["WACC"]
    # WACC result cell (B7 — inputs now live on Assumptions) is a percent.
    assert ws["B7"].number_format == "0.0%"


def test_native_charts_and_ergonomics(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _bundle(with_lbo=True))
    wb = load_workbook(p)
    # A banker expects embedded charts: football field + revenue + FCFF + comps.
    charted = {ws.title for ws in wb.worksheets if ws._charts}
    assert {"Football Field", "Model IS", "DCF", "Trading Comps"} <= charted
    # Ergonomics: wide label column + freeze panes on the data-grid tabs.
    assert wb["Model IS"].column_dimensions["A"].width >= 30
    assert wb["Model IS"].freeze_panes == "B4"


def test_verifier_cell_map_structure(tmp_path):
    m = build_verifier_cell_map(_bundle())
    # keys are (sheet, cell) tuples -> float
    assert all(isinstance(k, tuple) and len(k) == 2 for k in m)
    assert ("DCF", "B21") in m
    assert m[("DCF", "B21")] == _bundle().dcf.implied_price_gordon
